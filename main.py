# -*- coding: utf-8 -*-
"""
Created on Fri Mar 13 19:58:34 2020

@author: p_aru
"""

import numpy as np
import casadi as cad
import matplotlib.pyplot as plt
from utilities import dynamicmodel, dynsim, insilicoexp, dynsen, lrf, lrtest, biofmodel
from utilities_for_NLP import construct_NLP_collocation
from openpyxl import Workbook
import pyomo.environ as py
import pyomo.dae as pyd

prng = 10
np.random.seed(prng)


ubT = 18 # final time
N = 18 # number of control intervals
nsp = 12
mg = (ubT/N)
sigma_y = np.array([[0.01, 0], [0, 0.05]])
nm = []
nm += [nsp]
x_0 = [2.4, 0.001]
lbx_0 = [1.0, 0.001]
ubx_0 = [10.0, 0.001]

xp_0 = [0.0] * (2 * 4)

lbx = [0, 0]  # [-0.25, -inf]
ubx = [np.inf, np.inf]

lbxp = [-np.inf] * 8
ubxp = [np.inf] * 8

lbtheta = [0.589,  0.018, 1.045,  0.005]
ubtheta = lbtheta

lbu = [0.05, 5]
ubu = [0.20, 35]

truetheta = [0.31, 0.18, 0.55, 0.05]

f, n_x, n_u, n_theta = dynamicmodel()

problem, w0, lbw, ubw, lbg, ubg, trajectories, discrete = \
    construct_NLP_collocation(f, x_0, xp_0, lbx_0, ubx_0, lbx, ubx, lbu, ubu, lbtheta, ubtheta, lbxp, ubxp, ubT, N, nsp, mg)


#solver = cad.nlpsol('solver', 'bonmin', problem, {"discrete": discrete})#'ipopt', problem) # MINLP
solver = cad.nlpsol('solver', 'ipopt', problem)#, {"discrete": discrete})#'ipopt', problem) # relaxed problem
# Function to get x and u trajectories from w

sol = solver(x0=w0, lbx=lbw, ubx=ubw, lbg=lbg, ubg=ubg)

w_opt = sol['x'].full().flatten()

x_opt, u_opt, xp_opt, T, b = trajectories(sol['x'])
x_opt = x_opt.full()# to numpy array
u_opt = u_opt.full()# to numpy array#
xp_opt = xp_opt.full()# to numpy array
T_opt = T.full()
b_opt = b.full()

x_pred = dynsim(x_opt[:,0], u_opt, lbtheta, 18, 18)
x_m = insilicoexp(x_opt[:,0], u_opt, truetheta, 18, 18, nm[-1], sigma_y, b_opt)
x_q = dynsen(x_opt[:,0], u_opt, lbtheta, 18, 18, b_opt, nm[-1])
lr = lrf(x_opt[:,0], u_opt, truetheta, 18, 18, nm[-1], sigma_y, b_opt, x_m)
a = lrtest(x_opt[:,0], u_opt, lbtheta, truetheta, 18, 18, nm[-1], sigma_y, b_opt, x_m)
# xp_r = np.reshape(xp_opt, (nx[0], ntheta[0], 1+N))


FIM_t = []# np.zeros([4,4])
FIM1_det = []
FIM1_trace = []
FIM_det = []
FIM_trace = []
FIM1_mineigv = []
FIM1_maxeigv = []
for i in range(np.shape(xp_opt)[1] - 1):
    xp_r = np.reshape(xp_opt[:,i+1], (2, 4))
#    vv = np.zeros([ntheta[0], ntheta[0], 1 + N])
#    for i in range(0, 1 + N):
    FIM_t += [xp_r.T @ np.linalg.inv(np.array([[0.01, 0], [0, 0.05]])) @ xp_r]# + np.linalg.inv(np.array([[0.01, 0, 0, 0], [0, 0.05, 0, 0], [0, 0, 1, 0], [0, 0, 0, 0.2]]))

obsFIM = []
for i in range(np.shape(FIM_t)[0]):
    obsFIM += [np.linalg.inv((((10.0 - 0.001)**2)/12) * np.identity(n_theta)) + sum(FIM_t[:i+1])]
    
for i in range(np.shape(xp_opt)[1] - 1):
    FIM1_det += [2 * cad.sum1(cad.log(cad.diag(cad.chol(obsFIM[i]))))]
    FIM1_trace += [cad.trace(obsFIM[i])]
    FIM_det += [2 * np.sum(np.log(np.diag(np.linalg.cholesky(obsFIM[i]))))]
    FIM_trace += [np.trace(obsFIM[i])]
    FIM1_mineigv += [min(np.linalg.eig(obsFIM[i])[0])]
    FIM1_maxeigv += [max(np.linalg.eig(obsFIM[i])[0])]
    


V_theta = np.linalg.inv(sum(FIM_t[i] for i in range(len(FIM_t))))
theta1 = w_opt[n_x + n_x * n_theta:
               n_x + n_x * n_theta + n_theta]
    
t_value = np.zeros(n_theta)
for i in range(n_theta):
    t_value[i] = theta1[i] / np.sqrt(V_theta[i, i])

FIM1_trace_array = np.array(FIM1_trace)
FIM1_det_array = np.array(FIM1_det)
FIM1_maxeigv_array = np.array(FIM1_maxeigv)

tmeas = np.where(b_opt > 0.99)[0][0:nm[-1]]+1
ig_theta = np.array([0.01,0.01,0.01,0.01])
lb_theta = np.array([1e-3,1e-3,1e-3,1e-3])
ub_theta = np.array([1e3,1e3,1e3,1e3])
solver = py.SolverFactory('ipopt')
optmodel = biofmodel(x_opt[:,0], u_opt, x_m.T, sigma_y, tmeas, ig_theta, lb_theta, ub_theta, T_opt[0][0], N)
soln = solver.solve(optmodel)
est = np.array([py.value(optmodel.theta[0]), py.value(optmodel.theta[1]), py.value(optmodel.theta[2]), py.value(optmodel.theta[3])])
objf = py.value(optmodel.objfun)

x_pred_act = dynsim(x_opt[:,0], u_opt, est, 18, 18)

#vv = np.zeros([ntheta[0], ntheta[0], 1+N])
#for i in range(0, 1+N):
#    vv[:, :, i] = xp_r[:, :, i].T @ np.array([[0.01, 0], [0, 0.05]]) @ xp_r[:, :, i]
#V = np.linalg.inv(sum(xp_r[:, :, i].T @ np.array([[0.01, 0], [0, 0.05]]) @ xp_r[:, :, i] for i in range(0, 1+N)))
#theta1 = w_opt[nx[0] + nx[0] * ntheta[0]:
#               nx[0] + nx[0] * ntheta[0] + ntheta[0] + 1]
#t = np.zeros(ntheta)
#for i in range(ntheta[0]): t[i] = theta1[i] / np.sqrt(V[i, i])
    
# Plot the result
tgrid = np.linspace(0, T_opt[0], N+1)
plt.figure(1)
plt.clf()
plt.plot(tgrid, x_opt[0], '--')
plt.plot(tgrid, x_opt[1], '-')
plt.step(tgrid, np.append(np.nan, u_opt[0]), '--')
plt.step(tgrid, np.append(np.nan, u_opt[1]), '-')
plt.xlabel('t')
plt.legend(['x1', 'x2', 'u1', 'u2'])
plt.grid()

plt.figure(2)
plt.clf()
plt.plot(np.linspace(1,T_opt[0],N), FIM1_trace, '-')
plt.plot(np.linspace(1,T_opt[0],N), FIM1_det, '--')
plt.plot(np.linspace(1,T_opt[0],N), FIM1_maxeigv, '-o')
plt.xlabel('t')
plt.legend(['FIM_tr', 'FIM_det', 'FIM_maxeigv'])
plt.grid()
plt.show()

# Saving the output as an excel file
wb = Workbook()
ws = wb.active
ws.title = 'MBDoE results'
ws.cell(row = 1, column = 1, value = ('Time'))
ws.cell(row = 1, column = 2, value = ('Optimal u1'))
ws.cell(row = 1, column = 3, value = ('Optimal u2'))
ws.cell(row = 1, column = 4, value = ('Optimal x1'))
ws.cell(row = 1, column = 5, value = ('Optimal x2'))
ws.cell(row = 1, column = 6, value = ('Optimal sampling'))
for i in range (N+1):
    ws.cell(row = i + 2, column = 1, value = (tgrid[i][0]))
for i in range(N):
    ws.cell(row = i + 2, column = 2, value = (u_opt[0][i]))
    ws.cell(row = i + 2, column = 3, value = (u_opt[1][i]))
ws.cell(row = 20 + 0, column = 2, value = (u_opt[0][-1]))
ws.cell(row = 20 + 0, column = 3, value = (u_opt[1][-1]))
for i in range(N+1):
    ws.cell(row = i + 2, column = 4, value = (x_opt[0][i]))
    ws.cell(row = i + 2, column = 5, value = (x_opt[1][i]))
for i in range(N):
    ws.cell(row = i + 3, column = 6, value = (b_opt[i][0]))
ws.cell(row = 0 + 2, column = 6, value = (0))
    
ws.cell(row = 1, column = 8, value = ('Time'))
ws.cell(row = 1, column = 9, value = ('FIM trace'))
ws.cell(row = 1, column = 10, value = ('FIM det'))
ws.cell(row = 1, column = 11, value = ('FIM maxeigv'))
ws.cell(row = 1, column = 12, value = ('Optimal sampling'))

ws.cell(row = 2, column = 9, value = (0))
ws.cell(row = 2, column = 10, value = (0))
ws.cell(row = 2, column = 11, value = (0))
ws.cell(row = 2, column = 12, value = (0))
for i in range (N+1):
    ws.cell(row = i + 2, column = 8, value = (tgrid[i][0]))
for i in range(N):
    ws.cell(row = i + 3, column = 9, value = (FIM1_trace_array[i]))
    ws.cell(row = i + 3, column = 10, value = (FIM1_det_array[i]))
    ws.cell(row = i + 3, column = 11, value = (FIM1_maxeigv_array[i]))
    ws.cell(row = i + 3, column = 12, value = (b_opt[i][0]))
    
ws.cell(row = 1, column = 13, value = ('Sampled time'))
ws.cell(row = 1, column = 14, value = ('x1m'))
ws.cell(row = 1, column = 15, value = ('x2m'))
ws.cell(row = 1, column = 16, value = ('Time'))
ws.cell(row = 1, column = 17, value = ('x1'))
ws.cell(row = 1, column = 18, value = ('x2'))
for i in range(np.shape(tmeas)[0]):
    ws.cell(row = i+2, column = 13, value = (tmeas[i]))
    ws.cell(row = i+2, column = 14, value = (x_m[:,0][i]))
    ws.cell(row = i+2, column = 15, value = (x_m[:,1][i]))
for i in range(np.shape(tgrid)[0]):
    ws.cell(row = i+2, column = 16, value = (tgrid[i][0]))
    ws.cell(row = i+2, column = 17, value = (x_pred_act[:,0][i]))
    ws.cell(row = i+2, column = 18, value = (x_pred_act[:,1][i]))
    
wb.save('MBDoEfull_resultspremiere.xlsx')

## Saving the output as a text file
#myfile = open('results.txt', 'w')
#myfile.write('results of MBDoE')
#myfile.close()